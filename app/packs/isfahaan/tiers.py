"""ISFAHAAN — Tiers Odoo ↔ Pennylane : moteur de synchronisation des CLIENTS (puis FOURNISSEURS).

Pourquoi : le connecteur Odoo × Pennylane crée un tiers Pennylane à chaque facture dont le
partenaire Odoo ne partage AUCUN identifiant avec un tiers existant. Règle confirmée par
Pennylane (Camélia Herbil, 03/09/2026) : rapprochement dès qu'UN identifiant est commun parmi
  · adresse email · numéro de compte tiers (code comptable) · SIREN / SIRET / n° de TVA
  · numéro de facture.
Sinon -> nouveau tiers (doublon). Corrections manuelles via la table d'équivalence de
Pennylane Connect ; fusion de tiers possible dans Pennylane (UI), PAS par API.

Ce moteur, en LECTURE (run_tiers_sync) :
  1. lit les fiches Odoo de tête (customer_rank/supplier_rank > 0) + leur activité (factures),
     les tiers Pennylane + leur compte (411/401) + l'activité réelle du compte (écritures) ;
  2. établit le SIREN de chaque fiche (registre/TVA Odoo, reg_no/TVA Pennylane, valeurs
     validées par le client, sinon proposition de l'annuaire officiel) et en déduit le n° de TVA ;
  3. SIMULE les critères du connecteur -> « identifiants communs » actuels ;
  4. choisit le tiers Pennylane CANONIQUE (celui qui porte l'historique) et liste les doublons ;
  5. prépare le plan : ce qu'il faut écrire côté Pennylane (reg_no, vat_number, emails,
     reference ODOO_<id>) et côté Odoo (vat, company_registry, email, ref) pour que le
     connecteur rapproche à coup sûr ; ce qui reste humain (fusions) est listé.
Puis, en ÉCRITURE séparée et tracée (run_tiers_apply), applique le plan. JAMAIS d'IBAN.
"""
import csv
import difflib
import io
import json
import re
import time
import unicodedata
from collections import defaultdict
from datetime import datetime, timedelta

import httpx
from sqlmodel import Session, select, delete

from app.core.db import engine
from app.core.connectors import odoo, pennylane
from app.models import Company, TiersMatch, Setting

_TZ = timedelta(hours=4)

KINDS = {
    "client": {"rank": "customer_rank", "moves": ["out_invoice", "out_refund"],
               "pl_list": "/customers", "pl_put": "/company_customers/{id}",
               "pl_post": "/company_customers", "acc": "411", "label": "clients"},
    "fournisseur": {"rank": "supplier_rank", "moves": ["in_invoice", "in_refund"],
                    "pl_list": "/suppliers", "pl_put": "/suppliers/{id}",
                    "pl_post": "/suppliers", "acc": "401", "label": "fournisseurs"},
}
_LEGAL = {"SARL", "SAS", "SASU", "EURL", "SA", "SCI", "SNC", "STE", "SOCIETE", "ETS", "GROUPE",
          "GISF", "GJ", "DE", "LA", "LE", "LES", "DU", "DES", "ST", "SAINT", "SAINTE", "ET"}


# ------------------------------------------------------------------ normalisations
def _clean(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    return "".join(c for c in s if not unicodedata.combining(c)).upper()


def _toks(s):
    return {t for t in re.split(r"[^A-Z0-9]+", _clean(s)) if t and t not in _LEGAL and len(t) > 1}


def _ids(*vals):
    """Identifiants fiscaux normalisés depuis N'IMPORTE QUEL champ : SIREN:, SIRET:, TVA:."""
    out = set()
    for v in vals:
        s = re.sub(r"[\s.\-]", "", str(v or "")).upper()
        if not s:
            continue
        m = re.match(r"^FR[0-9A-Z]{2}(\d{9})$", s)
        if m:
            out |= {"TVA:" + s, "SIREN:" + m.group(1)}
            continue
        d = re.sub(r"\D", "", s)
        if len(d) == 14:
            out |= {"SIRET:" + d, "SIREN:" + d[:9]}
        elif len(d) == 9:
            out.add("SIREN:" + d)
    return out


def _siren_of(ids):
    s = {i[6:] for i in ids if i.startswith("SIREN:")}
    return next(iter(s)) if len(s) == 1 else None


def _siret_of(ids):
    s = {i[6:] for i in ids if i.startswith("SIRET:")}
    return next(iter(s)) if len(s) == 1 else None


def _tva(siren):
    """N° de TVA intracommunautaire FR calculé (clé = (12 + 3 × (SIREN mod 97)) mod 97)."""
    if not siren or not re.fullmatch(r"\d{9}", siren):
        return None
    return f"FR{(12 + 3 * (int(siren) % 97)) % 97:02d}{siren}"


def _emails(v):
    if isinstance(v, list):
        return {str(e).strip().lower() for e in v if e and "@" in str(e)}
    return {e.strip().lower() for e in re.split(r"[;,\s]+", str(v or "")) if "@" in e}


def _jac(a, b):
    return len(a & b) / max(1, len(a | b))


# ------------------------------------------------------------------ annuaire officiel
def _annuaire(name, cache):
    """recherche-entreprises.api.gouv.fr -> (siren, nom officiel, ville, similarité) ou None."""
    if name in cache:
        return cache[name]
    variants = [name]
    m = re.match(r"^(.*?)\((.+?)\)\s*$", name or "")
    if m:
        variants = [m.group(2).strip(), m.group(1).strip(), name]
    best = None
    for q in variants:
        qq = " ".join(_toks(q)) or _clean(q)
        if len(qq) < 3:
            continue
        try:
            r = httpx.get("https://recherche-entreprises.api.gouv.fr/search",
                          params={"q": qq, "per_page": 5}, timeout=30)
            results = (r.json() or {}).get("results") or []
        except Exception:
            results = []
        time.sleep(0.35)
        for x in results:
            nom = x.get("nom_complet") or x.get("nom_raison_sociale") or ""
            simil = difflib.SequenceMatcher(None, " ".join(sorted(_toks(q))), " ".join(sorted(_toks(nom)))).ratio()
            cp = str((x.get("siege") or {}).get("code_postal") or "")
            ville = (x.get("siege") or {}).get("libelle_commune") or ""
            # clientèle réunionnaise : un homonyme hors 974 est presque toujours faux
            score = simil + (0.15 if cp.startswith("974") else -0.5)
            if x.get("etat_administratif") == "C":
                score -= 0.3
            if best is None or score > best[0]:
                best = (score, x.get("siren"), nom, f"{cp} {ville}".strip(), simil)
        if best and best[0] >= 0.95:
            break
    out = (best[1], best[2], best[3], best[4]) if (best and best[0] >= 0.60) else None
    cache[name] = out
    return out


# ------------------------------------------------------------------ lectures
def _pull_odoo(oc, K):
    fields = ["name", "vat", "ref", "email", "phone", "is_company", K["rank"], "active"]
    if oc.has_field("res.partner", "company_registry"):
        fields.append("company_registry")
    if oc.has_field("res.partner", "siret"):
        fields.append("siret")
    return oc.search_read("res.partner", [("parent_id", "=", False), (K["rank"], ">", 0)], fields)


def _odoo_activity(oc, K):
    dom = [("move_type", "in", K["moves"]), ("state", "=", "posted")]
    out = {}
    for extra, key in (([], "total"), ([("invoice_date", ">=", "2026-01-01")], "2026")):
        try:
            grp = oc.execute("account.move", "read_group", [dom + extra, ["id"], ["commercial_partner_id"]])
        except Exception:
            grp = []
        for x in grp:
            pid = (x.get("commercial_partner_id") or [None])[0]
            if pid:
                out.setdefault(pid, {})[key] = x.get("commercial_partner_id_count", 0)
    return out


def _pull_pl(pl, K):
    out, cur = [], None
    while True:
        params = {"limit": 100}
        if cur:
            params["cursor"] = cur
        d = pl.get(K["pl_list"], **params)
        out += d.get("items") or []
        if not d.get("has_more"):
            break
        cur = d.get("next_cursor")
    return out


def _pl_accounts(pl):
    acc, cur = {}, None
    while True:
        params = {"limit": 100}
        if cur:
            params["cursor"] = cur
        d = pl.get("/ledger_accounts", **params)
        for a in d.get("items") or []:
            acc[a["id"]] = a
        if not d.get("has_more"):
            break
        cur = d.get("next_cursor")
    return acc


def _pl_activity(pl, tiers, acc, ctx, base_step):
    act = {}
    for i, t in enumerate(tiers):
        la = (t.get("ledger_account") or {}).get("id")
        a = acc.get(la) or {}
        try:
            lines = pl.account_lines(la) if la else []
        except Exception:
            lines = []
        dates = sorted(l.get("date") for l in lines if l.get("date"))
        act[t["id"]] = {"account": a.get("number"), "n": len(lines),
                        "n26": sum(1 for x in dates if x >= "2026-01-01"),
                        "last": dates[-1] if dates else None,
                        "solde": round(sum(float(l.get("debit") or 0) - float(l.get("credit") or 0) for l in lines), 2)}
        if (i + 1) % 20 == 0:
            ctx.progress(base_step, None, step=f"activité Pennylane {i + 1}/{len(tiers)}…")
        time.sleep(0.05)
    return act


def _validated_sirens(company_code, kind):
    """SIREN validés par le client (Setting tiers:<kind>:siren_validated = JSON {odoo_id: siren})."""
    with Session(engine) as s:
        st = s.exec(select(Setting).where(Setting.company_code == company_code,
                                          Setting.key == f"tiers:{kind}:siren_validated")).first()
    try:
        return {int(k): re.sub(r"\D", "", str(v)) for k, v in json.loads(st.value).items()} if st else {}
    except Exception:
        return {}


# ------------------------------------------------------------------ moteur
def run_tiers_sync(ctx, company_code, kind="client"):
    K = KINDS[kind]
    with Session(engine) as s:
        company = s.exec(select(Company).where(Company.code == company_code)).first()
    oc = odoo.for_company(company_code)
    pl = pennylane.for_company(company_code)
    if not oc or not pl:
        raise RuntimeError("clés Odoo et Pennylane requises")

    ctx.progress(0, 6, step="lecture Odoo…")
    partners = _pull_odoo(oc, K)
    oact = _odoo_activity(oc, K)
    ctx.log(f"Odoo : {len(partners)} fiche(s) {K['label']} de tête")
    ctx.progress(1, 6, step="lecture Pennylane…")
    tiers = _pull_pl(pl, K)
    acc = _pl_accounts(pl)
    ctx.log(f"Pennylane : {len(tiers)} tiers {K['label']}")
    ctx.progress(2, 6, step="activité des comptes Pennylane…")
    pact = _pl_activity(pl, tiers, acc, ctx, 2)
    validated = _validated_sirens(company_code, kind)

    for t in tiers:
        t["_ids"] = _ids(t.get("reg_no"), t.get("vat_number"))
        t["_em"] = _emails(t.get("emails"))
        t["_tk"] = _toks(t.get("name"))
        t["_a"] = pact[t["id"]]
        t["_odoo"] = int(str(t.get("reference"))[5:]) if str(t.get("reference") or "").startswith("ODOO_") else None
    for p in partners:
        p["_ids"] = _ids(p.get("company_registry"), p.get("siret"), p.get("vat"))
        p["_em"] = _emails(p.get("email"))
        p["_tk"] = _toks(p.get("name"))
        p["_ref"] = str(p.get("ref") or "").strip()
        p["_a"] = oact.get(p["id"], {})

    ctx.progress(3, 6, step="matching + simulation du connecteur…")
    cache = {}
    rows, used = [], set()
    now = datetime.utcnow()
    siren_owner = defaultdict(list)          # SIREN -> fiches Odoo (doublons Odoo)

    def _collective(t):
        """compte collectif racine (411 « Clients », 401 « Fournisseurs ») : jamais un candidat."""
        return (t["_a"]["account"] or "") in (K["acc"], K["acc"] + "000000") or \
            _clean(t.get("name")).strip() in ("CLIENTS", "FOURNISSEURS")

    def candidates(p):
        out = {}
        for t in tiers:
            if _collective(t):
                continue
            via = set()
            if p["_em"] & t["_em"]:
                via.add("email")
            if p["_ids"] & t["_ids"]:
                via.add("fiscal")
            if p["_ref"] and p["_ref"] == (t["_a"]["account"] or ""):
                via.add("compte")
            if t["_odoo"] == p["id"]:
                via.add("ref_ODOO")
            if via:
                out[t["id"]] = (t, via)
                continue
            # nom seul : égalité stricte des mots, ou recouvrement ≥ 2/3 (évite « LA ZONE X » ≈ « LA ZONE Y »)
            if p["_tk"] and (p["_tk"] == t["_tk"] or _jac(p["_tk"], t["_tk"]) >= 0.67):
                out[t["id"]] = (t, set())
        return out

    def _drop_contradictions(cands, siren):
        """un candidat qui porte un AUTRE SIREN n'est pas un doublon : c'est une autre société."""
        if not siren:
            return cands, []
        kept, dropped = {}, []
        for tid, (t, via) in cands.items():
            s = _siren_of(t["_ids"])
            if s and s != siren:
                dropped.append((t, via))
            else:
                kept[tid] = (t, via)
        return kept, dropped

    prelim = []
    for p in partners:
        cands = candidates(p)
        # SIREN retenu : validé > Odoo > Pennylane (unique parmi les candidats) > annuaire
        src, siren, prop = None, None, None
        if p["id"] in validated:
            siren, src = validated[p["id"]], "validé"
        elif _siren_of(p["_ids"]):
            siren, src = _siren_of(p["_ids"]), "odoo"
        else:
            # hérité de Pennylane seulement depuis un candidat SÛR (identifiant commun ou nom identique)
            pls = {_siren_of(t["_ids"]) for t, via in cands.values()
                   if _siren_of(t["_ids"]) and (via or t["_tk"] == p["_tk"])}
            if len(pls) == 1:
                siren, src = next(iter(pls)), "pennylane"
            else:
                hit = _annuaire(p.get("name"), cache)
                if hit:
                    siren, src = hit[0], "annuaire"
                    prop = f"{hit[1]} · {hit[2]} · similarité {hit[3]:.0%}"
                else:
                    prop = "introuvable dans l'annuaire"
        if siren:
            siren_owner[siren].append(p)
        prelim.append((p, cands, siren, src, prop))

    for p, cands, siren, src, prop in prelim:
        siret = _siret_of(p["_ids"])
        tva = _tva(siren)
        a = p["_a"]
        cands, dropped = _drop_contradictions(cands, siren)
        contra = [t for t, via in dropped if via]          # même email/compte mais AUTRE SIREN : à signaler
        # canonique = celui qui porte l'historique (écritures 2026, puis total, puis ref connecteur)
        ordered = sorted(cands.values(), key=lambda tv: (tv[0]["_a"]["n26"], tv[0]["_a"]["n"],
                                                         1 if "ref_ODOO" in tv[1] else 0), reverse=True)
        canon, via = (ordered[0] if ordered else (None, set()))
        dups = [t for t, _ in ordered[1:]]
        for t, _ in ordered:
            used.add(t["id"])
        # conflit d'identité : rapproché par identifiant fiscal mais noms sans aucun mot commun
        conflict = bool(canon) and "fiscal" in via and p["_tk"] and canon["_tk"] and not (p["_tk"] & canon["_tk"])

        plan_pl, plan_od = {}, {}
        if canon:
            if siren and _siren_of(canon["_ids"]) != siren:
                plan_pl["reg_no"] = siren
            if tva and (canon.get("vat_number") or "").replace(" ", "").upper() != tva:
                plan_pl["vat_number"] = tva
            if p["_em"] - canon["_em"]:
                plan_pl["emails"] = sorted(canon["_em"] | p["_em"])
            if str(canon.get("reference") or "") != f"ODOO_{p['id']}":
                plan_pl["reference"] = f"ODOO_{p['id']}"
        elif siren:
            plan_pl["create"] = {"name": p.get("name"), "reg_no": siren, "vat_number": tva,
                                 "emails": sorted(p["_em"])}
        if tva and re.sub(r"\s", "", str(p.get("vat") or "")).upper() != tva:
            plan_od["vat"] = tva
        if "company_registry" in p and not re.sub(r"\D", "", str(p.get("company_registry") or "")) and (siret or siren):
            plan_od["company_registry"] = siret or siren
        if canon and not p["_em"] and canon["_em"]:
            plan_od["email"] = sorted(canon["_em"])[0]
        if canon and not p["_ref"] and canon["_a"]["account"]:
            plan_od["ref"] = canon["_a"]["account"]

        # statut / mode / actions
        if siren and len(siren_owner[siren]) > 1:
            others = ", ".join(x.get("name") for x in siren_owner[siren] if x["id"] != p["id"])
            status, mode = "doublon_odoo", "MANUEL"
            note = (f"même SIREN {siren} que : {others} — Pennylane (et la facture électronique) identifient le tiers "
                    "par le SIREN : un client = une société, les magasins = adresses de livraison")
            act_u = ("CONFIRMER que ce sont des établissements de la même société, puis FUSIONNER ces fiches Odoo "
                     "(Contacts > Fusionner, garder la plus complète) ; chaque magasin devient une adresse de livraison")
            act_ia = "Après la fusion : j'équipe la fiche survivante (plan Pennylane/Odoo) et je re-contrôle"
        elif contra and not canon:
            status, mode = "conflit_identifiant", "MANUEL"
            note = "email/compte commun avec « " + contra[0].get("name") + f" » mais SIREN différent ({_siren_of(contra[0]['_ids'])} vs {siren})"
            act_u = "VÉRIFIER : quelle fiche porte le bon SIREN ? corriger dans Odoo ou Pennylane"
            act_ia = "Rien tant que ce n'est pas tranché"
        elif conflict:
            status, mode = "conflit_identifiant", "MANUEL"
            note = f"même identifiant fiscal mais noms sans rapport : « {canon.get('name')} » (compte {canon['_a']['account']})"
            act_u = "VÉRIFIER : même société ? sinon corriger le SIREN/TVA erroné dans Odoo ou Pennylane"
            act_ia = "Rien tant que ce n'est pas tranché"
        elif not canon:
            near = max((t for t in tiers if t["id"] not in used and not _collective(t)),
                       key=lambda t: _jac(p["_tk"], t["_tk"]), default=None)
            nj = _jac(p["_tk"], near["_tk"]) if near else 0
            if near and nj >= 0.3 and not (_siren_of(near["_ids"]) and siren and _siren_of(near["_ids"]) != siren):
                status, mode = "absent_pennylane", "A_VALIDER"
                note = f"aucun identifiant commun ; tiers Pennylane PROCHE : « {near.get('name')} » (compte {near['_a']['account']}, {near['_a']['n']} écr.)"
                act_u = f"TRANCHER : « {near.get('name')} » est-il ce client ? OUI = je l'équipe (SIREN {siren or '?'}) ; NON = je crée un tiers"
                act_ia = "Après réponse : équipe le tiers existant OU crée le tiers Pennylane"
            elif siren and src in ("odoo", "validé", "pennylane"):
                status, mode = "absent_pennylane", "AUTO"
                note = "aucun tiers Pennylane ne partage un identifiant ni un nom proche"
                act_u = "Rien (NON si ce client ne doit plus être facturé)"
                act_ia = f"CRÉE le tiers Pennylane (nom, SIREN {siren}, TVA {tva}" + (", email" if p["_em"] else "") + ")"
            elif siren:
                status, mode = "absent_pennylane", "A_VALIDER"
                note = "aucun tiers Pennylane proche ; SIREN proposé par l'annuaire"
                act_u = f"VALIDER le SIREN {siren} ({prop})"
                act_ia = "Après validation : écrit le SIREN/TVA dans Odoo et CRÉE le tiers Pennylane"
            else:
                status, mode = "sans_siren", "SAISIE"
                note = "aucun tiers Pennylane proche ; SIREN introuvable"
                act_u = "RECHERCHER le SIREN (annuaire-entreprises.data.gouv.fr) et me le donner"
                act_ia = "Ensuite : écrit SIREN/TVA dans Odoo et crée le tiers Pennylane"
        else:
            human = []
            real_dups = [t for t in dups if t["_a"]["n"] > 0]
            empty_dups = [t for t in dups if t["_a"]["n"] == 0]
            if real_dups:
                human.append("FUSIONNER dans Pennylane (Tiers > fusionner) les doublons listés vers le canonique « "
                             f"{canon.get('name')} » (compte {canon['_a']['account']})")
            if empty_dups:
                human.append("ARCHIVER dans Pennylane les doublons VIDES listés (aucune écriture)")
            if not via:
                note = "rapproché par le NOM seulement — AUCUN identifiant commun : le connecteur créerait un doublon"
            else:
                note = "identifiants communs : " + ", ".join(sorted(via))
            if real_dups or empty_dups:
                status = "doublon_pennylane"
            elif not via:
                status = "sans_identifiant_commun"
            else:
                status = "ok"
            if src == "annuaire":
                mode = "A_VALIDER"
                act_u = f"VALIDER le SIREN {siren} ({prop})" + (" ; puis " + " ; ".join(human) if human else "")
                act_ia = "Après validation : " + _plan_text(plan_pl, plan_od)
            elif human:
                mode = "MANUEL"
                act_u = " ; ".join(human)
                act_ia = _plan_text(plan_pl, plan_od) + " (fait aussi avant la fusion : sans risque)"
            elif not siren:
                mode = "SAISIE"
                act_u = "RECHERCHER le SIREN et me le donner (le rapprochement ne tient qu'au nom" + (
                    "/email)" if via else ")")
                act_ia = _plan_text(plan_pl, plan_od) or "Ensuite : écrit SIREN/TVA des deux côtés"
            elif plan_pl or plan_od:
                mode = "AUTO"
                act_u = "Rien à faire"
                act_ia = _plan_text(plan_pl, plan_od)
            else:
                mode = "RIEN"
                act_u = "Rien à faire"
                act_ia = "—"

        rows.append(TiersMatch(
            company_id=company.id, kind=kind, odoo_id=p["id"], odoo_name=p.get("name"),
            odoo_ref=p["_ref"] or None, odoo_vat=p.get("vat") or None,
            odoo_registry=(p.get("company_registry") or p.get("siret") or None),
            odoo_email=", ".join(sorted(p["_em"])) or None,
            odoo_inv_total=int(a.get("total") or 0), odoo_inv_2026=int(a.get("2026") or 0),
            siren=siren, siren_source=src, siret=siret, tva=tva, annuaire=prop,
            pl_id=canon["id"] if canon else None, pl_name=canon.get("name") if canon else None,
            pl_account=canon["_a"]["account"] if canon else None,
            pl_reg_no=(canon.get("reg_no") or None) if canon else None,
            pl_vat=(canon.get("vat_number") or None) if canon else None,
            pl_emails=", ".join(sorted(canon["_em"])) if canon else None,
            pl_reference=(canon.get("reference") or None) if canon else None,
            pl_lines=canon["_a"]["n"] if canon else 0, pl_lines_2026=canon["_a"]["n26"] if canon else 0,
            pl_solde=canon["_a"]["solde"] if canon else None,
            pl_dups=" | ".join(f"#{t['id']} {t.get('name')} · cpte {t['_a']['account']} · {t['_a']['n']} écr."
                               for t in dups) or None,
            match_via=", ".join(sorted(via)) or None, status=status, mode=mode,
            action_user=act_u, action_ia=act_ia,
            plan_pl=json.dumps(plan_pl, ensure_ascii=False) if plan_pl else None,
            plan_odoo=json.dumps(plan_od, ensure_ascii=False) if plan_od else None,
            last_synced=now))

    # orphelins Pennylane (aucune fiche Odoo ne les atteint)
    for t in sorted(tiers, key=lambda x: (x.get("name") or "").upper()):
        if t["id"] in used:
            continue
        best = max(partners, key=lambda p: _jac(p["_tk"], t["_tk"]), default=None)
        j = _jac(best["_tk"], t["_tk"]) if best else 0
        note = (f"proche de la fiche Odoo « {best.get('name')} » (similarité {j:.0%})" if j >= 0.34
                else "aucune fiche Odoo proche")
        if t["_a"]["n"] == 0:
            act_u = "ARCHIVER dans Pennylane (aucune écriture)"
        elif t["_a"]["n26"] == 0:
            act_u = "Historique clos (aucune écriture 2026) : laisser tel quel, ou fusionner vers le tiers actif"
        else:
            act_u = "VÉRIFIER : encore mouvementé en 2026 sans fiche Odoo — quel client Odoo facture dessus ?"
        rows.append(TiersMatch(
            company_id=company.id, kind=kind, odoo_id=None,
            pl_id=t["id"], pl_name=t.get("name"), pl_account=t["_a"]["account"],
            pl_reg_no=t.get("reg_no") or None, pl_vat=t.get("vat_number") or None,
            pl_emails=", ".join(sorted(t["_em"])) or None, pl_reference=t.get("reference") or None,
            pl_lines=t["_a"]["n"], pl_lines_2026=t["_a"]["n26"], pl_solde=t["_a"]["solde"],
            siren=_siren_of(t["_ids"]), status="orphelin_pennylane", mode="MANUEL",
            action_user=act_u, action_ia="Rien — je re-contrôle après", last_synced=now,
            annuaire=note))

    # Excel + compteurs AVANT la persistance (les objets sont détachés/expirés après commit)
    ctx.progress(4, 6, step="Excel…")
    stamp = (datetime.utcnow() + _TZ)
    xlsx = _excel(rows, company, kind, stamp)
    counts = defaultdict(int)
    modes = defaultdict(int)
    for r in rows:
        counts[r.status] += 1
        modes[r.mode] += 1
    no_id = sum(1 for r in rows if r.odoo_id and r.pl_id and not r.match_via)

    ctx.progress(5, 6, step="enregistrement…")
    with Session(engine) as s:
        s.exec(delete(TiersMatch).where(TiersMatch.company_id == company.id, TiersMatch.kind == kind))
        for r in rows:
            s.add(r)
        s.commit()
    ctx.add_artifact("xlsx", f"{stamp.strftime('%Y%m%d %H%M')} cadrage_{K['label']}_odoo_pennylane_{company_code} T{ctx.run_id}.xlsx",
                     xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    L = [f"TIERS ODOO ↔ PENNYLANE — {K['label'].upper()} — {company.name} — tâche #{ctx.run_id} du {stamp.strftime('%d/%m/%Y %H:%M')}", "",
         f"  Fiches Odoo : {len(partners)} · tiers Pennylane : {len(tiers)}",
         f"  Fiches Odoo SANS identifiant commun avec Pennylane (doublon garanti par le connecteur) : {no_id}", "",
         "== STATUTS =="] + [f"  {k:<26}: {v}" for k, v in sorted(counts.items())] + \
        ["", "== MODES =="] + [f"  {k:<12}: {v}" for k, v in sorted(modes.items())] + \
        ["", "Plan : Excel joint (une ligne par fiche, colonnes ACTION UTILISATEUR / ACTION IA). Aucune écriture n'a été faite."]
    ctx.set_report("\n".join(L))
    return (f"⚠️ Tiers {K['label']} — {counts.get('ok', 0)} ok · {modes.get('AUTO', 0)} AUTO · "
            f"{modes.get('A_VALIDER', 0)} à valider · {modes.get('MANUEL', 0)} manuels · {no_id} sans identifiant commun")


def _plan_text(plan_pl, plan_od):
    parts = []
    if plan_pl:
        if "create" in plan_pl:
            parts.append("Pennylane : CRÉE le tiers")
        else:
            parts.append("Pennylane : écrit " + ", ".join(f"{k}={v if not isinstance(v, list) else '+'.join(v)}"
                                                          for k, v in plan_pl.items()))
    if plan_od:
        parts.append("Odoo : écrit " + ", ".join(f"{k}={v}" for k, v in plan_od.items()))
    return " ; ".join(parts)


# ------------------------------------------------------------------ Excel
def _excel(rows, company, kind, stamp):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    MODES = {"RIEN": ("🟢 Rien à faire", "D8F0DE"), "AUTO": ("🔵 AUTO (Vaelan écrit)", "CFE2F8"),
             "A_VALIDER": ("🟡 À VALIDER (SIREN annuaire)", "FFF3BF"), "SAISIE": ("🟠 SIREN à rechercher", "FCE8CC"),
             "MANUEL": ("🔴 MANUEL (fusion)", "F8D2D5")}
    LBL = {"ok": "OK", "sans_identifiant_commun": "Sans identifiant commun", "doublon_pennylane": "Doublon Pennylane",
           "doublon_odoo": "Doublon Odoo", "conflit_identifiant": "Conflit d'identifiant",
           "absent_pennylane": "Absent de Pennylane", "sans_siren": "Sans SIREN", "orphelin_pennylane": "Orphelin Pennylane"}
    white = Font(bold=True, color="FFFFFF"); hdr = PatternFill("solid", fgColor="1F2430")
    thin = Border(bottom=Side(style="thin", color="DDDDDD")); bold = Font(bold=True)
    wb = Workbook(); ws = wb.active; ws.title = "Synthèse"
    for col, w in (("A", 34), ("B", 10), ("C", 100)):
        ws.column_dimensions[col].width = w
    ws["A1"] = f"CADRAGE {KINDS[kind]['label'].upper()} ODOO ↔ PENNYLANE — {company.name}"; ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"généré le {stamp.strftime('%d/%m/%Y %H:%M')} (Réunion) · lecture seule · règle du connecteur Pennylane : rapprochement dès qu'UN identifiant est commun (email, compte tiers, SIREN/SIRET/TVA)"
    ws["A2"].font = Font(color="777777")
    ws["A4"] = ("MODE D'EMPLOI : colonne MODE = qui fait quoi. Les lignes AUTO sont écrites par Vaelan (Pennylane : reg_no, TVA, "
                "emails, référence ODOO_<id> ; Odoo : TVA, registre, email, réf.) sur ordre. Les lignes À VALIDER attendent ton OUI "
                "sur le SIREN proposé. MANUEL = fusion de fiches (impossible par API). Jamais d'IBAN.")
    ws["A4"].font = Font(bold=True, color="8A6D00")
    r = 6
    mc = defaultdict(int); sc = defaultdict(int)
    for x in rows:
        mc[x.mode] += 1; sc[x.status] += 1
    for m, (lbl, color) in MODES.items():
        c = ws.cell(row=r, column=1, value=lbl); c.fill = PatternFill("solid", fgColor=color)
        ws.cell(row=r, column=2, value=mc.get(m, 0)); r += 1
    r += 1
    for st, lbl in LBL.items():
        if sc.get(st):
            ws.cell(row=r, column=1, value=lbl); ws.cell(row=r, column=2, value=sc[st]); r += 1

    HEAD = ["Statut", "MODE", "ID Odoo", "Fiche Odoo", "Réf. Odoo", "TVA Odoo", "Registre Odoo", "Email Odoo",
            "Fact. total", "Fact. 2026", "SIREN retenu", "Source", "TVA calculée", "Proposition annuaire",
            "ID Pennylane", "Tiers Pennylane", "Compte", "reg_no PL", "TVA PL", "Emails PL", "Réf. PL",
            "Écritures", "Écr. 2026", "Solde", "Identifiants communs", "Doublons Pennylane", "Constat",
            "ACTION UTILISATEUR", "ACTION IA", "Validation (OUI/NON)"]
    WID = [22, 22, 8, 30, 10, 15, 15, 26, 8, 8, 11, 10, 14, 40, 13, 30, 11, 11, 14, 26, 11, 8, 8, 11, 20, 46, 46, 52, 52, 14]

    def sheet(name, data):
        w = wb.create_sheet(name)
        for i, (h, wd) in enumerate(zip(HEAD, WID), 1):
            c = w.cell(row=1, column=i, value=h); c.font = white; c.fill = hdr
            w.column_dimensions[get_column_letter(i)].width = wd
        for j, x in enumerate(data, 2):
            vals = [LBL.get(x.status, x.status), MODES.get(x.mode, (x.mode, ""))[0], x.odoo_id, x.odoo_name, x.odoo_ref,
                    x.odoo_vat, x.odoo_registry, x.odoo_email, x.odoo_inv_total, x.odoo_inv_2026, x.siren,
                    x.siren_source, x.tva, x.annuaire if x.odoo_id else None, x.pl_id, x.pl_name, x.pl_account, x.pl_reg_no,
                    x.pl_vat, x.pl_emails, x.pl_reference, x.pl_lines, x.pl_lines_2026, x.pl_solde, x.match_via,
                    x.pl_dups, (x.annuaire if not x.odoo_id else None) or "", x.action_user, x.action_ia, ""]
            for i, v in enumerate(vals, 1):
                c = w.cell(row=j, column=i, value=v); c.border = thin
                if i == 2:
                    c.fill = PatternFill("solid", fgColor=MODES.get(x.mode, ("", "FFFFFF"))[1]); c.font = bold
                if i == 28:
                    c.fill = PatternFill("solid", fgColor=MODES.get(x.mode, ("", "FFFFFF"))[1])
                if i == 30:
                    c.fill = PatternFill("solid", fgColor="FFF3BF")
                c.alignment = Alignment(vertical="top", wrap_text=(i in (14, 26, 27, 28, 29)))
        w.freeze_panes = "E2"
        w.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{len(data) + 1}"

    order = {"MANUEL": 0, "A_VALIDER": 1, "SAISIE": 2, "AUTO": 3, "RIEN": 4}
    sheet("Correspondance", sorted([x for x in rows if x.odoo_id], key=lambda x: (order.get(x.mode, 9), -x.odoo_inv_2026)))
    sheet("Orphelins Pennylane", sorted([x for x in rows if not x.odoo_id], key=lambda x: -x.pl_lines_2026))
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------------ application (écriture)
def run_tiers_apply(ctx, company_code, kind="client", target="pennylane"):
    """Applique le plan des lignes AUTO (et MANUEL/A_VALIDER pour la partie identifiants déjà sûre :
    source SIREN = odoo / pennylane / validé). Jamais billing_iban. Audit CSV."""
    K = KINDS[kind]
    with Session(engine) as s:
        company = s.exec(select(Company).where(Company.code == company_code)).first()
        rows = s.exec(select(TiersMatch).where(TiersMatch.company_id == company.id, TiersMatch.kind == kind)).all()
    oc = odoo.for_company(company_code)
    pl = pennylane.for_company(company_code)
    audit = [["cible", "id", "nom", "champs", "resultat"]]
    done = errs = skipped = 0

    def _applicable(r):
        """Seules les lignes SÛRES sont écrites : identifiants d'un tiers rapproché sans ambiguïté
        (ok / sans identifiant commun / canonique d'un doublon Pennylane) et créations AUTO.
        Jamais : SIREN à valider ou à rechercher, doublons Odoo (fiche survivante inconnue), conflits."""
        if not r.odoo_id or r.mode in ("A_VALIDER", "SAISIE") or r.siren_source == "annuaire":
            return False
        if r.status in ("conflit_identifiant", "doublon_odoo", "sans_siren", "orphelin_pennylane"):
            return False
        if r.status == "absent_pennylane":
            return r.mode == "AUTO"
        return r.status in ("ok", "sans_identifiant_commun", "doublon_pennylane")

    for i, r in enumerate(rows):
        if not _applicable(r):
            continue
        plan = json.loads((r.plan_pl if target == "pennylane" else r.plan_odoo) or "{}")
        if not plan:
            continue
        ctx.progress(i, len(rows), step=f"{target} {r.odoo_name}…")
        try:
            if target == "pennylane":
                if "create" in plan:
                    body = {k: v for k, v in plan["create"].items() if v}
                    body["reference"] = f"ODOO_{r.odoo_id}"
                    code, resp = pl.send_json("POST", K["pl_post"], body)
                else:
                    body = {k: v for k, v in plan.items() if k != "billing_iban"}
                    code, resp = pl.send_json("PUT", K["pl_put"].format(id=r.pl_id), body)
                ok = 200 <= code < 300
                audit.append(["pennylane", r.pl_id or "création", r.pl_name or r.odoo_name, json.dumps(body, ensure_ascii=False),
                              "ok" if ok else f"HTTP {code} {str(resp)[:120]}"])
            else:
                body = {k: v for k, v in plan.items() if k in ("vat", "company_registry", "email", "ref")}
                oc.execute("res.partner", "write", [[r.odoo_id], body])
                ok = True
                audit.append(["odoo", r.odoo_id, r.odoo_name, json.dumps(body, ensure_ascii=False), "ok"])
            if ok:
                done += 1
                with Session(engine) as s:
                    x = s.get(TiersMatch, r.id)
                    if target == "pennylane":
                        x.applied_pl_at = datetime.utcnow()
                    else:
                        x.applied_odoo_at = datetime.utcnow()
                    s.add(x); s.commit()
            else:
                errs += 1
        except Exception as e:
            errs += 1
            audit.append([target, r.pl_id if target == "pennylane" else r.odoo_id, r.odoo_name, json.dumps(plan, ensure_ascii=False), str(e)[:150]])
        time.sleep(0.2)
    buf = io.StringIO(); csv.writer(buf, delimiter=";").writerows(audit)
    stamp = datetime.utcnow() + _TZ
    ctx.add_artifact("csv", f"{stamp.strftime('%Y%m%d %H%M')} audit_tiers_{target}_{company_code} T{ctx.run_id}.csv",
                     buf.getvalue().encode("utf-8-sig"), "text/csv")
    ctx.set_report(f"APPLICATION {target.upper()} — {K['label']} — {company.name} — {stamp.strftime('%d/%m/%Y %H:%M')}\n"
                   f"  écrit : {done} · erreurs : {errs}\n  audit CSV joint (une ligne par écriture).")
    return f"{'✅' if not errs else '⚠️'} Tiers {K['label']} → {target} : {done} écrit(s), {errs} erreur(s)"
